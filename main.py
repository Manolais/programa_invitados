import smtplib as sm
import os
from openpyxl import Workbook,load_workbook
import sys


def identTerminal():
    import sys
    p = sys.platform
    if p == 'aix':
        return
    elif p == 'linux':
        return 'clear'
    elif p == 'win32':
        return 'cls'
    elif p == 'cygwin':
        return
    elif p == 'darwin':
        return 'clear'


personas = []

while True:
    if not os.path.exists('personas.xlsx'):
        wb = Workbook('personasxlsx')
        sheet1 = wb.create_sheet('Sheet1')
        wb.remove_sheet('Sheet')
        sheet1['A1'] = 'Nombre'
        sheet1['B1'] = 'Teléfono'
        sheet1['C1'] = 'Email'
        wb.save('personas.xlsx')
    os.system(identTerminal())
    print('1 - Ingreso')
    print('2 - Ver cantidad de personas')
    print('3 - Salir y enviar')
    n = input('>>> ')
    if n == '1':
        while True:
            os.system(identTerminal())
            per = input('Ingrese el nombre: ')
            tel = input('Ingrese número telefónico: ')
            try:
                tel = int(tel)
                break
            except ValueError:
                print('Ingrese un valor numérico')
                input('Presione enter para continuar...')
        mail = input('Ingrese el mail: ')
        personas.append({'Nombre': per,'Telefono': tel,'Email': mail})
        print('¡Invitado ingresado correctamente!')
        input('Presione enter para continuar...')

    elif n == '2':
        os.system(identTerminal())
        print('1 - Ver invitados en ram')
        print('2 - Ver invitados en excel')
        p = input('>>>')
        if p == '1':
            if len(personas) == 0:
                os.system(identTerminal())
                print('Lista de invitados vacia, por favor agregue invitados para poder visualizarlos')
                input('Presione enter para continuar...')
            else:
                os.system(identTerminal())
                print('--------------------')
                for i in personas:
                    print('')
                    print('Invitado',personas.index(i)+1)
                    print()
                    print('Nombre:  ',i['Nombre'])
                    print('Teléfono:',i['Telefono'])
                    print('Email:   ',i['Email'])
                    print('--------------------')
                input('Presione enter para continuar...')
        if p == '2':
            wb = load_workbook('personas.xlsx')
            sheet1 = wb.get_sheet_by_name('Sheet1')
            sheet1['A1'] = 'Nombre'
            sheet1['B1'] = 'Teléfono'
            sheet1['C1'] = 'Email'
            if sheet1['A2'].value == None:
                os.system(identTerminal())
                print('Excel vacio, vuelva cuando guarde los cambios')
                input('Presione enter para continuar...')
            elif type(sheet1['A2'].value) == str:
                os.system(identTerminal())
                print('--------------------')
                for i in range(len(sheet1['A'])-1):
                    print('Invitado', i+1)
                    print()
                    print('Nombre:  ',sheet1['A'+str(i+2)].value)
                    print('Teléfono:',sheet1['B'+str(i+2)].value)
                    print('Email:   ',sheet1['C'+str(i+2)].value)
                    print('--------------------')
                input('Presione enter para continuar...')
        


    elif n == '3':
        wb = load_workbook('personas.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet['A1'] = 'Nombre'
        sheet['B1'] = 'Teléfono'
        sheet['C1'] = 'Email'
        for i in personas:
            sheet['A'+str(len(sheet['A'])+1)] = i['Nombre']
            sheet['B'+str(len(sheet['B']))] = i['Telefono']
            sheet['C'+str(len(sheet['C']))] = i['Email']
        wb.save('personas.xlsx')
        os.system(identTerminal())
        sv = sm.SMTP('smtp.gmail.com',587)
        sv.starttls()
        sv.login('programainvitados@gmail.com','prog_invit')
        sv.sendmail('programainvitados@gmail.com','bsasmanuelsilva@gmail.com',str(personas))
        os.system(identTerminal())
        print('Gracias por usar el programa, guardando invitados...')
        print('Correo enviado correctamente!')
        input('Presione enter para salir...')
        exit()
    elif n=='exit()':
        exit()
    else:
        print('Ingrese una opción valida')
        input('Presione enter para continuar...')